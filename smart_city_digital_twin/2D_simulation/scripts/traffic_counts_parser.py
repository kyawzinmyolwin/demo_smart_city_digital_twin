#!/usr/bin/env python3
# pyright: reportMissingModuleSource=false
"""
Parse Miovision-style traffic-count workbooks into one long CSV.

Flow: SCANNING → VALIDATING → VERIFYING → PARSING → CSV.

Usage:
- Run everything under `data/input/source_data/`:
  `python3 traffic_counts_parser.py`
- Run a specific workbook (path is relative to `data/input/source_data/`):
  `python3 traffic_counts_parser.py "2020 Intersection/I1234 Some Site 06-05-2026.xlsx"`

References:
- Christchurch City Council. (2025). *Intersection traffic counts database*. 
  https://ccc.govt.nz/transport/improving-our-transport-and-roads/traffic-count-data/intersection-traffic-counts-database 
- Christchurch City Council. (2025). *Road intersection (OpenData)*. 
  https://opendata-christchurchcity.hub.arcgis.com/datasets/4912c568d9a742caa630873278554932_6/explore 
- Traffic Engineering and Management Limited. (2025). *Intersection workbook folder* [Data set]. 
  https://drive.google.com/drive/folders/1oP5gcuKR1bHB9Xn2B4ILZZd_LhpMggxU 
- Land Information New Zealand. (2024). *Concord: Coordinate system concordance*. 
  https://www.geodesy.linz.govt.nz/concord/index.cgi 

Output: `traffic_DDMMMYYYY_hhmmss.csv` in ``data/output/demand/`` (NZ time). Columns include
approach_bound and destination_id / destination_latitude / destination_longitude (from
intersection_geo.csv octant neighbours). The daily ``traffic_DDMMMYYYY.log`` is written there too.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path


def _running_under_debugpy() -> bool:
    """True when the process was started by debugpy (VS Code / Cursor debugger)."""
    try:
        for a in getattr(sys, "orig_argv", ()):
            if "debugpy" in str(a).lower():
                return True
    except Exception:
        pass
    for a in sys.argv:
        if "debugpy" in str(a).lower():
            return True
    return False


def _project_venv_python_for_hint() -> Path | None:
    """First ``.venv`` Python binary found (repo root preferred), or ``None``."""
    script_path = Path(__file__).resolve()
    roots = (script_path.parent.parent, script_path.parent)
    if sys.platform == "win32":
        for root in roots:
            for name in ("python.exe", "python3.exe"):
                p = root / ".venv" / "Scripts" / name
                if p.is_file():
                    return p
    else:
        for root in roots:
            for name in ("python3", "python"):
                p = root / ".venv" / "bin" / name
                if p.is_file():
                    return p
    return None


def _hint_debugger_use_project_venv() -> None:
    """Tell users running under debugpy to select ``.venv`` (re-exec is disabled)."""
    if not _running_under_debugpy():
        return
    v = _project_venv_python_for_hint()
    if v is None:
        return
    sys.stderr.write(
        "\nDebugger note: switching to ``.venv`` automatically is disabled under "
        "debugpy (``os.execv`` would end the debug session).\n"
        f"  Select this interpreter: {v}\n"
        "  Cursor / VS Code: Command Palette → Python: Select Interpreter → "
        "``.venv/bin/python``, or start Run and Debug with the configuration "
        "``traffic_counts_parser (.venv)`` in ``.vscode/launch.json``.\n\n"
    )
    sys.stderr.flush()


def _maybe_reexec_with_project_venv() -> None:
    """
    If ``.venv`` exists beside this file or one folder up (repo root), and the
    current interpreter is not already that venv's Python, replace this process
    with the venv interpreter so dependencies install with ``pip`` match imports.

    Disabled when ``TRAFFIC_PARSER_NO_VENV_REEXEC`` is ``1``/``true``/``yes``,
    or when debugpy started this process (``os.execv`` drops the debugger).

    Skips re-exec when ``sys.prefix`` already matches that ``.venv`` (even if
    ``sys.executable`` resolves to the base Homebrew/framework binary).
    """
    if os.environ.get("TRAFFIC_PARSER_NO_VENV_REEXEC", "").strip().lower() in (
        "1",
        "true",
        "yes",
    ):
        return
    if _running_under_debugpy():
        return

    script_path = Path(__file__).resolve()
    # Prefer the repo-level ``.venv`` (parent of ``smart_city_digital_twin/``) over a venv
    # inside ``smart_city_digital_twin/`` so we do not pick an empty or stale local env first.
    search_roots = (script_path.parent.parent, script_path.parent)
    venv_py: Path | None = None
    if sys.platform == "win32":
        for root in search_roots:
            for name in ("python.exe", "python3.exe"):
                p = root / ".venv" / "Scripts" / name
                if p.is_file():
                    venv_py = p
                    break
            if venv_py is not None:
                break
    else:
        for root in search_roots:
            for name in ("python3", "python"):
                p = root / ".venv" / "bin" / name
                if p.is_file():
                    venv_py = p
                    break
            if venv_py is not None:
                break

    if venv_py is None:
        return

    venv_root = venv_py.parent.parent
    try:
        if Path(sys.prefix).resolve() == venv_root.resolve():
            return
    except OSError:
        pass

    # Use the venv launcher path (not ``Path.resolve()``): resolving follows the
    # symlink to the base interpreter and would drop venv ``sys.prefix`` context.
    argv = [str(venv_py), str(script_path), *sys.argv[1:]]
    os.execv(str(venv_py), argv)


_maybe_reexec_with_project_venv()

import argparse
import csv
import re
import shutil
import subprocess
import time
import warnings
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any
from zoneinfo import ZoneInfo

from _sim_root import DATA_INPUT_DIR, DATA_OUTPUT_DIR, DEMAND_DIR, SIM_ROOT

from pipeline_progress import _tick_refresh_due

# Quote non-numeric fields so values with spaces are delimited.
_CSV_WRITE_OPTS: dict[str, Any] = {"quoting": csv.QUOTE_NONNUMERIC, "lineterminator": "\n"}

# Generated CSV and run logs default to this script's directory (not the shell cwd).
_PARSER_APP_DIR = SIM_ROOT
_DEMAND_DIR = DEMAND_DIR

_RUN_LOG_FH: Any | None = None

_ANSI_CSI_RE = re.compile(r"\x1b\[[0-9;]*[A-Za-z]")


def _strip_ansi_for_log(s: str) -> str:
    """Remove terminal escape sequences so the run log stays readable."""
    out = _ANSI_SGR_RE.sub("", s)
    out = _ANSI_CSI_RE.sub("", out)
    return out


def _log_timestamp_nz() -> str:
    """Run-log timestamp, NZ local time, with milliseconds (e.g. ``[15:54:32.109]``)."""
    dt = datetime.now(ZoneInfo("Pacific/Auckland"))
    ms = dt.microsecond // 1000
    return f"[{dt.strftime('%H:%M:%S')}.{ms:03d}]"


def _run_log_line(message: str) -> None:
    """Write one timestamped line into the `.log` file (if enabled)."""
    if _RUN_LOG_FH is None:
        return
    text = message.rstrip("\r\n")
    clean = _strip_ansi_for_log(text)
    if not clean.strip():
        return
    _RUN_LOG_FH.write(f"{_log_timestamp_nz()} {clean}\n")
    _RUN_LOG_FH.flush()


def _stderr_line(message: str) -> None:
    """Write one normalised line to stderr and, if enabled, to the run log."""
    text = message.rstrip("\r\n")
    sys.stderr.write(f"{text}\n")
    _run_log_line(text)
    if sys.stderr.isatty():
        sys.stderr.flush()


def _plain_mode_file_line(
    show_ui: bool,
    *,
    ok: bool,
    ok_prefix: str,
    workbook: Path,
    survey_date: date | None = None,
) -> None:
    """
    Per-workbook status line.

    Non-TTY prints to stderr (and log). TTY writes to the log only.
    """
    short = _log_book_display_name(workbook, survey_date=survey_date)
    line = f"{ok_prefix}: {short}" if ok else f"FAILED: {short}"
    if show_ui:
        _run_log_line(line)
        return
    _stderr_line(line)


def _fmt_n(n: int) -> str:
    """Format an integer as #,##0."""
    return format(int(n), ",d")


def _fmt_n_block_rjust(
    n_scan: int, n_val: int, n_ver: int, n_row: int
) -> tuple[str, str, str, str]:
    """Comma-format the four totals and right-pad them to a shared width."""
    parts = [_fmt_n(n_scan), _fmt_n(n_val), _fmt_n(n_ver), _fmt_n(n_row)]
    w = max(len(p) for p in parts) if parts else 1
    return tuple(p.rjust(w) for p in parts)


def _emit_final_summary_block(
    *,
    tty_colours: bool,
    n_scan: int,
    n_val: int,
    n_ver: int,
    n_row: int,
) -> None:
    """Aligned four-line totals (SCANNED…PARSED)."""
    fs, fv, fver, fr = _fmt_n_block_rjust(n_scan, n_val, n_ver, n_row)
    lab_w = 11
    specs = (
        ("SCANNED:", fs, "files"),
        ("VALIDATED:", fv, "files"),
        ("VERIFIED:", fver, "files"),
        ("PARSED:", fr, "rows"),
    )
    for lab, num, unit in specs:
        if tty_colours:
            sys.stderr.write(
                f"\033[94m{lab:<{lab_w-1}}\033[0m \033[97m{num} {unit}\033[0m\033[K\n"
            )
        else:
            sys.stderr.write(f"{lab:<{lab_w}}{num} {unit}\n")
    sys.stderr.flush()


def _emit_tty_summary_line(
    label: str,
    count: int,
    unit: str,
    *,
    num_width: int = 7,
    label_width: int = 10,
) -> None:
    """TTY-only summary line."""
    if not sys.stderr.isatty():
        return
    num = _fmt_n(count).rjust(num_width)
    sys.stderr.write(
        f"\033[94m{label:<{label_width}}\033[0m \033[97m{num} {unit}\033[0m\033[K\n"
    )
    sys.stderr.flush()


def _pip_module_available(exe: str) -> bool:
    """True if ``python -m pip`` works for this interpreter."""
    r = subprocess.run(
        [exe, "-m", "pip", "--version"],
        capture_output=True,
        check=False,
    )
    return r.returncode == 0


def _deps_install_verified_in_this_process() -> bool:
    """
    After a subprocess ``pip install``, the running interpreter must reload import
    metadata so new site-packages entries are visible.

    Returns True only if ``import openpyxl`` succeeds in **this** process (same as
    ``sys.executable`` used to run this script).
    """
    import importlib

    importlib.invalidate_caches()
    try:
        import openpyxl  # noqa: F401
    except ImportError as err:
        sys.stderr.write(
            "pip finished, but this process still cannot import openpyxl.\n"
            f"  Interpreter: {sys.executable}\n"
            f"  Error: {err}\n"
            "Often Homebrew Python is PEP 668–managed (no installs into the base env), "
            "or the debugger stayed on that interpreter while ``.venv`` already has packages.\n"
            "Try in a terminal: python3 -m pip install openpyxl\n"
            "If you use Python 3.14 only outside a venv, wheels may be missing — "
            "prefer the project ``.venv`` or Python 3.12–3.13.\n"
        )
        sys.stderr.flush()
        _hint_debugger_use_project_venv()
        return False
    return True


def _ensurepip_bootstrap(exe: str) -> bool:
    """
    Install pip into this interpreter via the stdlib ``ensurepip`` module.

    Some installs (e.g. minimal Homebrew Python) ship without pip; ``-m pip`` then
    fails with ``No module named pip``.
    """
    sys.stderr.write(
        "pip is missing for this Python; running ensurepip (stdlib bootstrap)...\n"
    )
    sys.stderr.flush()
    for args in (
        [exe, "-m", "ensurepip", "--upgrade"],
        [exe, "-m", "ensurepip", "--upgrade", "--default-pip"],
    ):
        if subprocess.run(args, check=False).returncode == 0 and _pip_module_available(exe):
            return True
    return False


def _stderr_pep668_hint(stderr: str, req_file: Path) -> None:
    """Extra help when Homebrew / distro blocks ``pip install`` (PEP 668)."""
    if "externally-managed-environment" not in stderr:
        return
    sys.stderr.write(
        "\nHint (PEP 668): this interpreter blocks system-wide pip installs.\n"
        "  Create a project virtual environment and install there, then pick that\n"
        "  interpreter in your IDE (e.g. Cursor → Python: Select Interpreter):\n"
        "    python3 -m venv .venv\n"
        f"    .venv/bin/pip install -r \"{req_file.resolve()}\"\n\n"
    )
    sys.stderr.flush()


def _auto_install_requirements_once() -> bool:
    """
    Install parser deps via pip using this interpreter.

    Skipped when env ``TRAFFIC_PARSER_NO_AUTO_PIP`` is ``1``/``true``/``yes``.
    Uses ``requirements.txt`` beside this script if present; otherwise installs
    ``openpyxl`` only.

    If ``pip`` is absent, runs ``ensurepip`` once then retries ``pip install``.
    """
    if os.environ.get("TRAFFIC_PARSER_NO_AUTO_PIP", "").strip().lower() in (
        "1",
        "true",
        "yes",
    ):
        return False
    exe = sys.executable
    req = SIM_ROOT / "requirements.txt"
    if req.is_file():
        sys.stderr.write(f"Installing dependencies from {req.name} ({exe})...\n")
        cmd = [exe, "-m", "pip", "install", "-r", str(req)]
    else:
        sys.stderr.write(
            f"requirements.txt not beside script; installing openpyxl only ({exe})...\n"
        )
        cmd = [exe, "-m", "pip", "install", "openpyxl"]
    sys.stderr.flush()
    r = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if r.returncode == 0:
        return _deps_install_verified_in_this_process()
    if r.stderr:
        sys.stderr.write(r.stderr)
        sys.stderr.flush()
        _stderr_pep668_hint(r.stderr, req)
    if r.stdout:
        sys.stderr.write(r.stdout)
        sys.stderr.flush()
    if not _pip_module_available(exe):
        if _ensurepip_bootstrap(exe):
            r2 = subprocess.run(cmd, capture_output=True, text=True, check=False)
            if r2.returncode == 0:
                return _deps_install_verified_in_this_process()
            if r2.stderr:
                sys.stderr.write(r2.stderr)
                sys.stderr.flush()
                _stderr_pep668_hint(r2.stderr, req)
            if r2.stdout:
                sys.stderr.write(r2.stdout)
                sys.stderr.flush()
        sys.stderr.write(
            "ensurepip failed. Install pip for this interpreter manually, "
            "or pick another Python in your IDE (one that already has pip).\n"
        )
        sys.stderr.flush()
    return False


try:
    import openpyxl

    try:
        from openpyxl.utils.datetime import from_excel as _excel_serial_to_datetime
    except ImportError:
        _excel_serial_to_datetime = None  # type: ignore[misc, assignment]
except ImportError as e:
    if _auto_install_requirements_once():
        try:
            import openpyxl

            try:
                from openpyxl.utils.datetime import from_excel as _excel_serial_to_datetime
            except ImportError:
                _excel_serial_to_datetime = None  # type: ignore[misc, assignment]
        except ImportError as e2:
            _stderr_line(
                "Missing dependency: openpyxl  (pip install openpyxl). "
                "Auto-install failed or set TRAFFIC_PARSER_NO_AUTO_PIP=1 and install manually."
            )
            raise SystemExit(1) from e2
    else:
        _stderr_line(
            "Missing dependency: openpyxl  (pip install openpyxl). "
            "Place requirements.txt next to this script for automatic install, "
            "or set TRAFFIC_PARSER_NO_AUTO_PIP=1 to skip."
        )
        raise SystemExit(1) from e

try:
    import xlrd  # type: ignore[import-untyped]
except ImportError:
    xlrd = None  # type: ignore[misc, assignment]


def _xml_local_tag(tag: str) -> str:
    return tag.split("}")[-1] if tag.startswith("{") else tag


def _xml_attr_int(elem: Any, local: str) -> int | None:
    """First attribute whose local name matches (e.g. Index)."""
    for k, v in elem.attrib.items():
        if _xml_local_tag(k).lower() == local.lower():
            try:
                return int(v)
            except ValueError:
                return None
    return None


def _xml_cell_data_text(data_elem: Any) -> tuple[Any, str | None]:
    """Return (python_value, raw_type)."""
    typ = None
    for k, v in data_elem.attrib.items():
        if _xml_local_tag(k).lower() == "type":
            typ = v
            break
    text = (data_elem.text or "").strip()
    if typ is None:
        return (text if text else None, typ)
    tlow = typ.lower()
    if tlow == "number":
        try:
            return (float(text), typ)
        except ValueError:
            return (text, typ)
    if tlow == "boolean":
        return (text.lower() in ("1", "true", "yes"), typ)
    if tlow in ("datetime", "date"):
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return (datetime.strptime(text.split(".")[0], fmt), typ)
            except ValueError:
                continue
        return (text, typ)
    return (text if text else None, typ)


class _SimpleMergedRange:
    __slots__ = ("min_row", "max_row", "min_col", "max_col")

    def __init__(self, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col


def _parse_xml_spreadsheet_table(
    table_elem: Any,
) -> tuple[dict[tuple[int, int], Any], int, int, list[_SimpleMergedRange]]:
    """Sparse 1-based (row, col) -> value map for SpreadsheetML, including merged ranges."""
    cells: dict[tuple[int, int], Any] = {}
    merged: list[_SimpleMergedRange] = []
    max_r = max_c = 1
    row_num: int | None = None
    for row in table_elem:
        if _xml_local_tag(row.tag) != "Row":
            continue
        ri = _xml_attr_int(row, "Index")
        if ri is not None:
            row_num = ri
        else:
            row_num = 1 if row_num is None else row_num + 1
        col_num: int | None = None
        for cell in row:
            if _xml_local_tag(cell.tag) != "Cell":
                continue
            ci = _xml_attr_int(cell, "Index")
            if ci is not None:
                col_num = ci
            else:
                col_num = 1 if col_num is None else col_num + 1
            val: Any = None
            for child in cell:
                if _xml_local_tag(child.tag) != "Data":
                    continue
                val, _ = _xml_cell_data_text(child)
                break
            merge_across = _xml_attr_int(cell, "MergeAcross") or 0
            merge_down = _xml_attr_int(cell, "MergeDown") or 0
            cells[(row_num, col_num)] = val
            if merge_across or merge_down:
                merged.append(
                    _SimpleMergedRange(
                        row_num,
                        row_num + merge_down,
                        col_num,
                        col_num + merge_across,
                    )
                )
            max_r = max(max_r, row_num + merge_down)
            max_c = max(max_c, col_num + merge_across)
            # SpreadsheetML may omit the covered cells; advance the pointer so following
            # cells land in the correct column when Index is not present.
            if merge_across:
                col_num += merge_across
    return cells, max_r, max_c, merged


class _XmlSpreadsheetSheetAdapter:
    """Minimal worksheet API for Excel 2003 XML (SpreadsheetML)."""

    __slots__ = ("_cells", "_max_r", "_max_c", "merged_cells")

    def __init__(
        self,
        cells: dict[tuple[int, int], Any],
        max_r: int,
        max_c: int,
        merged_ranges: list[_SimpleMergedRange] | None = None,
    ) -> None:
        self._cells = cells
        self._max_r = max_r
        self._max_c = max_c
        self.merged_cells = type("MC", (), {"ranges": list(merged_ranges or [])})()

    @property
    def max_row(self) -> int:
        return max(1, self._max_r)

    @property
    def max_column(self) -> int:
        return max(1, self._max_c)

    def cell(self, row: int, col: int) -> Any:
        class _Cell:
            __slots__ = ("value",)

        o = _Cell()
        o.value = self._cells.get((row, col))
        return o


class XmlSpreadsheetBookAdapter:
    """Workbook adapter for SpreadsheetML."""

    __slots__ = ("sheetnames", "_sheets", "_by_index")

    def __init__(self, path: Path) -> None:
        tree = ET.parse(path)
        root = tree.getroot()
        self.sheetnames: list[str] = []
        sheets: list[_XmlSpreadsheetSheetAdapter] = []
        for child in root:
            if _xml_local_tag(child.tag) != "Worksheet":
                continue
            name = ""
            for k, v in child.attrib.items():
                if _xml_local_tag(k).lower() == "name":
                    name = v
                    break
            self.sheetnames.append(name or f"Sheet{len(self.sheetnames) + 1}")
            tbl = None
            for sub in child.iter():
                if _xml_local_tag(sub.tag) == "Table":
                    tbl = sub
                    break
            if tbl is None:
                cells, mr, mc, merged = {}, 1, 1, []
            else:
                cells, mr, mc, merged = _parse_xml_spreadsheet_table(tbl)
            sheets.append(_XmlSpreadsheetSheetAdapter(cells, mr, mc, merged))
        if not self.sheetnames:
            self.sheetnames = ["Sheet1"]
            sheets = [_XmlSpreadsheetSheetAdapter({}, 1, 1, [])]
        self._sheets = sheets
        self._by_index = sheets

    @property
    def worksheets(self) -> list[Any]:
        return self._sheets

    @property
    def active(self) -> Any:
        return self._sheets[0]

    def __getitem__(self, name: str) -> Any:
        try:
            i = self.sheetnames.index(name)
        except ValueError as e:
            raise KeyError(name) from e
        return self._sheets[i]

    def close(self) -> None:
        pass


class _XlrdMergedRange:
    __slots__ = ("min_row", "max_row", "min_col", "max_col")

    def __init__(self, rlo: int, rhi: int, clo: int, chi: int) -> None:
        self.min_row = rlo + 1
        self.max_row = rhi + 1
        self.min_col = clo + 1
        self.max_col = chi + 1


class _XlrdMergedCells:
    __slots__ = ("ranges",)

    def __init__(self, ranges: list[_XlrdMergedRange]) -> None:
        self.ranges = ranges


class XlrdSheetAdapter:
    """Worksheet adapter for xlrd."""

    __slots__ = ("_book", "_sheet", "merged_cells")

    def __init__(self, book: Any, sheet: Any) -> None:
        self._book = book
        self._sheet = sheet
        rngs: list[_XlrdMergedRange] = []
        for crange in sheet.merged_cells:
            rlo, rhi, clo, chi = crange
            rngs.append(_XlrdMergedRange(rlo, rhi, clo, chi))
        self.merged_cells = _XlrdMergedCells(rngs)

    @property
    def max_row(self) -> int:
        return max(1, self._sheet.nrows)

    @property
    def max_column(self) -> int:
        return max(1, self._sheet.ncols)

    def cell(self, row: int, col: int) -> Any:
        class _Cell:
            __slots__ = ("value",)

        o = _Cell()
        rowx, colx = row - 1, col - 1
        if rowx < 0 or colx < 0 or rowx >= self._sheet.nrows or colx >= self._sheet.ncols:
            o.value = None
            return o
        o.value = _xlrd_cell_to_value(self._book, self._sheet, rowx, colx)
        return o


class XlrdBookAdapter:
    """Workbook adapter for xlrd."""

    __slots__ = ("_book", "sheetnames", "_sheets")

    def __init__(self, book: Any) -> None:
        self._book = book
        self.sheetnames = list(book.sheet_names())
        self._sheets = [XlrdSheetAdapter(book, book.sheet_by_name(n)) for n in self.sheetnames]

    @property
    def worksheets(self) -> list[Any]:
        return self._sheets

    @property
    def active(self) -> Any:
        return self._sheets[0]

    def __getitem__(self, name: str) -> Any:
        return XlrdSheetAdapter(self._book, self._book.sheet_by_name(name))

    def close(self) -> None:
        self._book.release_resources()


def _xlrd_cell_to_value(book: Any, sheet: Any, rowx: int, colx: int) -> Any:
    assert xlrd is not None
    typ = sheet.cell_type(rowx, colx)
    val = sheet.cell_value(rowx, colx)
    if typ == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(val, book.datemode)
        except Exception:
            return val
    if typ == xlrd.XL_CELL_BOOLEAN:
        return bool(val)
    if typ in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if typ == xlrd.XL_CELL_ERROR:
        return None
    return val


def _peek_file_kind(path: Path) -> str:
    """
    Distinguish OOXML zip, OLE BIFF .xls, and SpreadsheetML XML.
    Returns: 'ooxml' | 'xls_biff' | 'spreadsheetml'
    """
    with path.open("rb") as f:
        head = f.read(8192)
    if len(head) >= 4 and head[:2] == b"PK":
        return "ooxml"
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls_biff"
    h = head.lstrip(b"\xef\xbb\xbf")
    if h.startswith(b"<?xml") or (h.startswith(b"<") and b"Workbook" in head[:200]):
        return "spreadsheetml"
    return "xls_biff"


def open_workbook_auto(path: Path) -> Any:
    """Open a supported workbook; call close_workbook()."""
    kind = _peek_file_kind(path)
    if kind == "ooxml":
        return openpyxl.load_workbook(path, data_only=True)
    if kind == "spreadsheetml":
        return XmlSpreadsheetBookAdapter(path)
    if xlrd is None:
        _stderr_line(
            "Missing dependency: xlrd is required for binary .xls files (pip install xlrd)."
        )
        raise SystemExit(1)
    try:
        book = xlrd.open_workbook(str(path), formatting_info=False)
    except Exception as e:
        _stderr_line(f"Could not open as legacy .xls ({path.name}): {e}")
        raise SystemExit(1) from e
    return XlrdBookAdapter(book)


def close_workbook(wb: Any) -> None:
    """Close any adapter workbook."""
    if hasattr(wb, "close"):
        wb.close()


def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _cell_value_follow_merged(ws: Any, row: int, col: int) -> Any:
    """
    Return the cell value, following merged ranges back to the first (top-left) cell.

    This matters because many Excel writers store the text only in the first cell of a merged
    area, while the other cells in the merged range appear empty when read programmatically.
    """
    v = ws.cell(row, col).value
    if v is not None:
        return v
    merged = getattr(ws, "merged_cells", None)
    ranges = getattr(merged, "ranges", None) if merged is not None else None
    if not ranges:
        return None
    for r in ranges:
        # openpyxl cell ranges and our adapters both provide min/max row/col.
        if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
            return ws.cell(r.min_row, r.min_col).value
    return None


def _format_time(v: Any) -> str:
    """Format an Excel time cell as HH:MM (zero-padded 24-hour)."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M")
    s = _norm_str(v)
    if not s:
        return ""
    s_up = s.strip().upper()
    # 12-hour with optional seconds, e.g. "1:30 PM", "12:00:00 AM".
    m12 = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?\s*(AM|PM)\s*$", s_up)
    if m12:
        h, mi = int(m12.group(1)), int(m12.group(2))
        if not (1 <= h <= 12 and 0 <= mi <= 59):
            return ""
        ap = m12.group(3)
        if ap == "PM" and h != 12:
            h += 12
        elif ap == "AM" and h == 12:
            h = 0
        return f"{h:02d}:{mi:02d}"
    # Plain H:MM or HH:MM (24-hour); optional seconds ignored.
    m24 = re.fullmatch(r"(\d{1,2}):(\d{2})(?::\d{2})?", s_up.replace(" ", ""))
    if m24:
        h, mi = int(m24.group(1)), int(m24.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return f"{h:02d}:{mi:02d}"
    try:
        f = float(v)
        if 0 <= f < 1:
            total_min = int(round(f * 24 * 60))
            h, m = divmod(total_min, 60)
            return f"{h:02d}:{m:02d}"
    except (TypeError, ValueError):
        pass
    return ""


def _extract_intersection_label(wb: Any, path: Path) -> str:
    """Best-effort intersection label."""
    if "Input" in wb.sheetnames:
        ws = wb["Input"]
        for r in range(1, min(40, (ws.max_row or 0) + 1)):
            key = _norm_str(ws.cell(r, 2).value).lower().rstrip(":")
            if key == "intersection":
                v = _norm_str(ws.cell(r, 3).value)
                if v:
                    return v
    roles = _miovision_metric_sheet_roles(wb)
    for key in _MIOVISION_LAYOUT_ROLE_ORDER:
        if key not in roles:
            continue
        ws = wb[roles[key]]
        for r in range(1, min(15, (ws.max_row or 0) + 1)):
            if _norm_str(ws.cell(r, 1).value).lower() == "study name":
                v = _norm_str(ws.cell(r, 2).value)
                if v:
                    return v
    return path.stem


def _extract_intersection_id_from_filename(path: Path) -> str | None:
    """Parse I#### from a filename."""
    m = re.match(r"(I\d+)[_\s.-]", path.name, re.I)
    if m:
        return m.group(1).upper()
    m = re.search(r"\b(I\d+)\b", path.stem, re.I)
    if m:
        return m.group(1).upper()
    return None


def _extract_intersection_id_from_label(label: str) -> str | None:
    m = re.search(r"\b(I\d+)\b", label, re.I)
    return m.group(1).upper() if m else None


def _compute_intersection_id(path: Path, wb: Any) -> str:
    lbl = _extract_intersection_label(wb, path)
    return (
        _extract_intersection_id_from_filename(path)
        or _extract_intersection_id_from_label(lbl)
        or ""
    )


def _iso_survey_date(iso: str) -> date | None:
    """Parse YYYY-MM-DD from intersection scan metadata ISO string."""
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso).date()
    except ValueError:
        return None


def _cell_to_date(v: Any) -> date | None:
    """Parse a date cell."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and _excel_serial_to_datetime is not None:
        try:
            dt = _excel_serial_to_datetime(float(v))
            if hasattr(dt, "date"):
                return dt.date()
            return dt
        except (ValueError, TypeError, OSError):
            pass
    # Excel serial without openpyxl helper.
    if isinstance(v, (int, float)) and 30000 < float(v) < 60000:
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=float(v))).date()
        except (ValueError, OSError):
            pass
    s = _norm_str(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _extract_survey_date_from_workbook(wb: Any) -> date | None:
    """Read survey date from Input sheet (Date:)."""
    if "Input" not in wb.sheetnames:
        return None
    ws = wb["Input"]
    for r in range(1, min(45, (ws.max_row or 0) + 1)):
        key = _norm_str(ws.cell(r, 2).value).lower().rstrip(":")
        if key == "date":
            raw = ws.cell(r, 3).value
            return _cell_to_date(raw)
    return None


def _extract_survey_date_from_filename(path: Path) -> date | None:
    """Parse trailing _DD-MM-YYYY or _MM-DD-YYYY from filename."""
    m = re.search(r"_(\d{2})-(\d{2})-(\d{4})(?:\.[^.]+)?$", path.name)
    if not m:
        return None
    a, b, yyyy = (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # Prefer DD-MM-YYYY; if impossible, fall back to MM-DD-YYYY.
    try:
        return date(yyyy, b, a)
    except ValueError:
        pass
    try:
        return date(yyyy, a, b)
    except ValueError:
        return None


def _log_book_display_name(
    path: Path, *, survey_date: date | None = None
) -> str:
    """
    Short label for per-file log lines: I####…DD-MM-YYYY.ext
    (intersection id from the name, date from filename tail or optional survey_date, extension lowercased).
    """
    iid = _extract_intersection_id_from_filename(path)
    head = iid if iid else "I?"
    d = _extract_survey_date_from_filename(path)
    if d is None and survey_date is not None:
        d = survey_date
    dpart = d.strftime("%d-%m-%Y") if d is not None else "?"
    ext = path.suffix.lower() if path.suffix else ""
    return f"{head}…{dpart}{ext}"


def _compute_survey_date(wb: Any, path: Path) -> date | None:
    return _extract_survey_date_from_workbook(wb) or _extract_survey_date_from_filename(
        path
    )


def _norm_sheet_key(name: str) -> str:
    """Lowercase alphanumeric-only key for fuzzy sheet-name matching."""
    return re.sub(r"[^a-z0-9]+", "", _norm_str(name).lower())


# Prefer likely vehicle-movement sheets for the layout grid.
_MIOVISION_LAYOUT_ROLE_ORDER: tuple[str, ...] = (
    "lights",
    "other_vehicles",
    "bicycles_on_road",
    "bicycles_crosswalk",
    "bicycles_generic",
    "totals",
)


def _infer_period_from_clock_time(time_str: str) -> str:
    """Derive AM/PM label from a clock string (12-hour with suffix or HH:MM 24-hour)."""
    u = time_str.strip().upper()
    if " AM" in f" {u}" or u.endswith("AM"):
        return "AM"
    if " PM" in f" {u}" or u.endswith("PM"):
        return "PM"
    m = re.fullmatch(r"(\d{2}):(\d{2})", u)
    if m:
        h = int(m.group(1))
        if 0 <= h < 12:
            return "AM"
        if 12 <= h <= 23:
            return "PM"
    return ""


def _miovision_metric_sheet_roles(wb: Any) -> dict[str, str]:
    """Map metric role -> sheet name (Miovision exports)."""
    role_to_name: dict[str, str] = {}
    for n in wb.sheetnames:
        k = _norm_sheet_key(n)
        if not k:
            continue
        low = n.strip().lower()
        if "total" in k and "sub" not in low:
            role_to_name.setdefault("totals", n)
        elif "other" in k and "vehicle" in k:
            role_to_name.setdefault("other_vehicles", n)
        elif "bicycle" in k and "road" in k:
            role_to_name.setdefault("bicycles_on_road", n)
        elif "bicycle" in k and "crosswalk" in k:
            role_to_name.setdefault("bicycles_crosswalk", n)
        elif k == "bicycles" or low == "bicycles":
            role_to_name.setdefault("bicycles_generic", n)
        elif "light" in k and "bicycle" not in k and "pedestrian" not in k:
            role_to_name.setdefault("lights", n)
    return role_to_name


def _pick_layout_worksheet(
    wb: Any, roles: dict[str, str]
) -> tuple[str | None, Any | None]:
    """First metric sheet with a Start Time movement grid."""
    for key in _MIOVISION_LAYOUT_ROLE_ORDER:
        if key not in roles:
            continue
        ws = wb[roles[key]]
        if _find_start_time_header_row(ws) is not None:
            return (key, ws)
    return (None, None)


def _find_start_time_header_row(ws: Any, *, scan: int = 200) -> int | None:
    """1-based row where column A is Start Time (movement grid)."""
    limit = min(scan, ws.max_row or 0)
    last: int | None = None
    for r in range(1, limit + 1):
        if _norm_str(ws.cell(r, 1).value).lower() == "start time":
            last = r
    return last


def _cell_int_count(ws: Any | None, row: int, col: int) -> int:
    if ws is None:
        return 0
    v = ws.cell(row, col).value
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


_ALLOWED_MOVEMENTS: frozenset[str] = frozenset(
    {
        "Bear Left",
        "Bear Right",
        "Hard Left",
        "Hard Right",
        "Left",
        "Right",
        "Thru",
        "U-Turn",
    }
)


def _is_allowed_movement(label: str) -> bool:
    return label in _ALLOWED_MOVEMENTS


def parse_miovision_metric_workbook(
    wb: Any,
    *,
    roles: dict[str, str],
) -> list[dict[str, Any]]:
    """Parse Miovision metric workbook (Start Time × movement grid)."""
    _lk, ws_layout = _pick_layout_worksheet(wb, roles)
    if ws_layout is None:
        raise ValueError(
            'Miovision layout: no metric sheet with "Start Time" in column A for a movement grid.'
        )
    hr = _find_start_time_header_row(ws_layout)
    if hr is None:
        raise ValueError(
            'Miovision layout: could not find "Start Time" in column A on the layout sheet.'
        )
    ha = hr - 2
    hd = hr - 1
    if ha < 1 or hd < 1:
        raise ValueError("Miovision layout: not enough header rows above Start Time row.")

    ws_lights = wb[roles["lights"]] if "lights" in roles else None
    ws_o = wb[roles["other_vehicles"]] if "other_vehicles" in roles else None
    ws_br = wb[roles["bicycles_on_road"]] if "bicycles_on_road" in roles else None
    ws_bgen = wb[roles["bicycles_generic"]] if "bicycles_generic" in roles else None
    ws_bcw = wb[roles["bicycles_crosswalk"]] if "bicycles_crosswalk" in roles else None
    ws_t = wb[roles["totals"]] if "totals" in roles else None

    max_col = ws_layout.max_column or 0
    max_row = ws_layout.max_row or 0
    start_r = hr + 1
    rows_out: list[dict[str, Any]] = []

    for r in range(start_r, max_row + 1):
        time_str = _format_time(ws_layout.cell(r, 1).value)
        if not time_str:
            continue
        period = _infer_period_from_clock_time(time_str)
        for col in range(2, max_col + 1):
            mov_label = _norm_str(ws_layout.cell(hr, col).value)
            if not mov_label or mov_label.lower() == "start time":
                continue
            if not _is_allowed_movement(mov_label):
                continue
            block = (col - 2) // 8
            base = 2 + block * 8
            approach_main = _norm_str(_cell_value_follow_merged(ws_layout, ha, base))
            direction_lbl = _norm_str(_cell_value_follow_merged(ws_layout, hd, col))
            # Keep direction information in separate columns for analysis.
            approach_header_1 = approach_main or direction_lbl
            approach_header_2 = direction_lbl if approach_main else ""
            mi = (col - 2) + 1
            bikes_road = (
                _cell_int_count(ws_br, r, col)
                if ws_br is not None
                else _cell_int_count(ws_bgen, r, col)
            )
            lights = _cell_int_count(ws_lights, r, col)
            other_vehicles = _cell_int_count(ws_o, r, col)
            bicycles_crosswalk = _cell_int_count(ws_bcw, r, col)
            totals = _cell_int_count(ws_t, r, col)
            if not (lights or other_vehicles or bikes_road or bicycles_crosswalk or totals):
                continue
            rows_out.append(
                {
                    "period": period,
                    "time": time_str,
                    "movement_index": mi,
                    "approach_header_1": approach_header_1,
                    "approach_header_2": approach_header_2,
                    "approach_bound": direction_lbl,
                    "movement": mov_label,
                    "lights": lights,
                    "other_vehicles": other_vehicles,
                    "bicycles_on_road": bikes_road,
                    "bicycles_crosswalk": bicycles_crosswalk,
                    "totals": totals,
                }
            )

    return rows_out


def _pip_style_progress_line(
    index: int,
    total: int,
    *,
    rows_so_far: int | None = None,
    files_so_far: int | None = None,
    mode: str = "parse",
    bar_width: int = 36,
    color: bool | None = None,
) -> str:
    """Legacy bar helper (parse mode); kept for reference — same bar fragments as live progress."""
    if color is None:
        color = sys.stderr.isatty()
    if total <= 0:
        frac = 1.0
    else:
        frac = min(1.0, index / total)
    filled = int(round(bar_width * frac))
    filled = min(max(filled, 0), bar_width)
    empty = bar_width - filled
    pct = 100.0 * frac
    pct_s = f"{pct:,.1f}%"
    if color:
        bar_s = (
            f"\033[36m{'━' * filled}\033[0m"
            f"\033[90m{'─' * empty}\033[0m"
        )
        pct_s = f"\033[33m{pct_s}\033[0m"
    else:
        bar_s = "━" * filled + "─" * empty
    if mode == "verify":
        suffix = f" ({_fmt_n(index)} files)"
    elif rows_so_far is not None:
        n_files = files_so_far if files_so_far is not None else index
        suffix = f" ({_fmt_n(rows_so_far)} rows from {_fmt_n(n_files)} files)"
    else:
        suffix = ""
    return f"         {bar_s}  {pct_s}{suffix}"


def _bar_and_pct_fragments(
    index: int,
    total: int,
    *,
    bar_width: int,
    color: bool | None = None,
) -> tuple[str, str]:
    """Filled bar string and percentage string (colour when TTY)."""
    if color is None:
        color = sys.stderr.isatty()
    if total <= 0:
        frac = 1.0
    else:
        frac = min(1.0, index / total)
    filled = int(round(bar_width * frac))
    filled = min(max(filled, 0), bar_width)
    empty = bar_width - filled
    pct = 100.0 * frac
    pct_s = f"{pct:,.1f}%"
    if color:
        bar_s = (
            f"\033[36m{'━' * filled}\033[0m"
            f"\033[90m{'─' * empty}\033[0m"
        )
        pct_s = f"\033[33m{pct_s}\033[0m"
    else:
        bar_s = "━" * filled + "─" * empty
    return bar_s, pct_s


_ANSI_SGR_RE = re.compile(r"\x1b\[[0-9;]*m")


def _visible_char_len(s: str) -> int:
    """Length without ANSI SGR sequences (wrap-safe width estimate)."""
    return len(_ANSI_SGR_RE.sub("", s))


def _stderr_progress_width() -> int:
    """Maximum width for each progress row (avoid terminal wrapping during redraw)."""
    try:
        cols = shutil.get_terminal_size((80, 24)).columns
    except OSError:
        cols = 80
    return max(48, cols - 2)


def _fmt_compact_rows(n: int) -> str:
    """Short row counts for narrow terminals."""
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 10_000:
        return f"{n / 1000:.1f}k"
    return _fmt_n(n)

def _pad_current_to_total_width(current_s: str, total_s: str) -> str:
    """Pad the current count so it lines up with the total's digit width."""
    return current_s.rjust(len(total_s))


def _compact_progress_suffix(
    action: str,
    index: int,
    total: int,
    rows_so_far: int | None,
) -> str:
    """Shorter parenthetical when the bar row would wrap (narrow terminal)."""
    if action == "PARSING":
        return f" ({_fmt_compact_rows(rows_so_far or 0)})"
    tot_s = _fmt_compact_rows(total)
    cur_s = _pad_current_to_total_width(_fmt_compact_rows(index), tot_s)
    return f" ({cur_s}/{tot_s})"


def _truncate_visible_plain(s: str, max_width: int) -> str:
    """Last resort: unstyled text truncated so it cannot wrap as multiple rows."""
    plain = _ANSI_SGR_RE.sub("", s)
    if len(plain) <= max_width:
        return s
    cut = max(1, max_width - 1)
    return plain[:cut] + "…"


def _ellipsis_end_filename(name: str, budget: int) -> str:
    """Shorten basename for progress row 1 without stripping colours elsewhere."""
    if budget <= 0:
        return ""
    if len(name) <= budget:
        return name
    if budget <= 1:
        return "…"[:budget]
    return name[: budget - 1] + "…"


def _two_phase_progress_rows(
    action: str,
    filename: str,
    index: int,
    total: int,
    *,
    rows_so_far: int | None = None,
    bar_width: int = 36,
    max_width: int | None = None,
) -> tuple[str, str]:
    """
    Row 1: ``action: filename``; row 2: ``bar NNN.N% (current/total)``.

    The basename is end-ellipsised only if wider than the terminal.
    """
    bar_opts = (bar_width, 32, 28, 22, 18, 14, 12, 10, 8)
    modes = (False, True)

    fn_row1 = filename
    if max_width is not None:
        prefix_len = len(f"{action}: ")
        room = max_width - prefix_len
        if room < 1:
            fn_row1 = "…"
        elif len(filename) > room:
            fn_row1 = _ellipsis_end_filename(filename, room)

    def rows(bar_w: int, compact: bool) -> tuple[str, str]:
        bw = min(bar_w, bar_width)
        bar_s, pct_s = _bar_and_pct_fragments(
            index, total, bar_width=bw, color=None
        )
        if compact:
            tot_s = _fmt_compact_rows(total)
            cur_s = _pad_current_to_total_width(_fmt_compact_rows(index), tot_s)
        else:
            tot_s = _fmt_n(total)
            cur_s = _pad_current_to_total_width(_fmt_n(index), tot_s)
        suffix = f" ({cur_s}/{tot_s})"
        if sys.stderr.isatty():
            row1 = f"\033[93m{action}:\033[0m \033[97m{fn_row1}\033[0m"
        else:
            row1 = f"{action}: {fn_row1}"
        row2 = f"{bar_s} {pct_s}{suffix}"
        return row1, row2

    r1, r2 = rows(bar_width, False)
    if max_width is None:
        return r1, r2

    if (
        _visible_char_len(r1) <= max_width
        and _visible_char_len(r2) <= max_width
    ):
        return r1, r2

    for compact in modes:
        for bw in bar_opts:
            r1, r2 = rows(bw, compact)
            if (
                _visible_char_len(r1) <= max_width
                and _visible_char_len(r2) <= max_width
            ):
                return r1, r2

    r1, r2 = rows(8, True)
    if _visible_char_len(r2) > max_width:
        r2 = _truncate_visible_plain(r2, max_width)
    return r1, r2


def _colour_status_message(message: str, *, enable: bool | None = None) -> str:
    """Colour status prefixes on a TTY."""
    if enable is None:
        enable = sys.stderr.isatty()
    if not enable:
        return message
    if message.startswith("UPDATED:"):
        return f"\033[32mUPDATED:\033[0m{message[len('UPDATED:'):]}"
    if message.startswith("SKIPPED:"):
        return f"\033[31mSKIPPED:\033[0m{message[len('SKIPPED:'):]}"
    if message.startswith("COMPLETED:"):
        # Deep blue (256-colour): 19 is close to navy.
        return f"\033[38;5;19mCOMPLETED:\033[0m{message[len('COMPLETED:'):]}"
    return message


class _DualPhaseProgress:
    """Four phases: two stderr rows per tick (phase + file; bar, percent, counts), rewritten with `\\033[2A`."""

    __slots__ = (
        "_can_rewrite",
        "_last",
        "_last_paint_at",
        "_last_phase_key",
        "_step",
        "_term_cols",
    )

    def __init__(self) -> None:
        self._can_rewrite = False
        self._last: dict[str, Any] | None = None
        self._step: str = "scan"
        self._term_cols: int | None = None
        self._last_paint_at = 0.0
        self._last_phase_key = ""

    def _reset_throttle(self) -> None:
        self._last_paint_at = 0.0
        self._last_phase_key = ""

    def invalidate(self) -> None:
        """Disable rewrite until the next tick."""
        self._can_rewrite = False
        self._term_cols = None
        self._reset_throttle()

    def begin_parse_phase(self) -> None:
        self._can_rewrite = False
        self._last = None
        self._step = "parse"
        self._term_cols = None
        self._reset_throttle()

    def tick_scan(
        self,
        *,
        scan_index: int,
        scan_total: int,
        scan_file: str,
    ) -> None:
        self._step = "scan"
        self._last = {
            "scan_index": scan_index,
            "scan_total": scan_total,
            "scan_file": scan_file,
        }
        self._paint()

    def tick_validate(
        self,
        *,
        validate_index: int,
        validate_total: int,
        validate_file: str,
    ) -> None:
        self._step = "validate"
        self._last = {
            "validate_index": validate_index,
            "validate_total": validate_total,
            "validate_file": validate_file,
        }
        self._paint()

    def tick_verify(
        self,
        *,
        verify_index: int,
        verify_total: int,
        verify_file: str,
    ) -> None:
        self._step = "verify"
        self._last = {
            "verify_index": verify_index,
            "verify_total": verify_total,
            "verify_file": verify_file,
        }
        self._paint()

    def tick_parse(
        self,
        *,
        parse_index: int,
        parse_total: int,
        parse_file: str,
        rows_so_far: int,
    ) -> None:
        self._step = "parse"
        self._last = {
            "parse_index": parse_index,
            "parse_total": parse_total,
            "parse_file": parse_file,
            "rows_so_far": rows_so_far,
        }
        self._paint()

    def _compose_active_rows(self, *, max_width: int | None = None) -> tuple[str, str]:
        lv = self._last
        assert lv is not None
        if self._step == "scan":
            return _two_phase_progress_rows(
                "SCANNING",
                lv["scan_file"],
                lv["scan_index"],
                lv["scan_total"],
                max_width=max_width,
            )
        if self._step == "validate":
            return _two_phase_progress_rows(
                "VALIDATING",
                lv["validate_file"],
                lv["validate_index"],
                lv["validate_total"],
                max_width=max_width,
            )
        if self._step == "verify":
            return _two_phase_progress_rows(
                "VERIFYING",
                lv["verify_file"],
                lv["verify_index"],
                lv["verify_total"],
                max_width=max_width,
            )
        return _two_phase_progress_rows(
            "PARSING",
            lv["parse_file"] or "—",
            lv["parse_index"],
            lv["parse_total"],
            rows_so_far=lv["rows_so_far"],
            max_width=max_width,
        )

    def _paint(self, *, force: bool = False) -> None:
        if self._last is None:
            return
        lv = self._last
        if self._step == "scan":
            index, total = lv["scan_index"], lv["scan_total"]
        elif self._step == "validate":
            index, total = lv["validate_index"], lv["validate_total"]
        elif self._step == "verify":
            index, total = lv["verify_index"], lv["verify_total"]
        else:
            index, total = lv["parse_index"], lv["parse_total"]
        if not _tick_refresh_due(
            index,
            total,
            self._last_paint_at,
            self._step,
            self._last_phase_key,
            force=force,
        ):
            return
        self._last_phase_key = self._step
        self._last_paint_at = time.perf_counter()
        mw: int | None = None
        if sys.stderr.isatty():
            if self._term_cols is None:
                self._term_cols = _stderr_progress_width()
            mw = self._term_cols
        r1, r2 = self._compose_active_rows(max_width=mw)
        body = f"{r1}\033[K\n{r2}\033[K\n"
        if self._can_rewrite:
            sys.stderr.write(f"\033[2A\r{body}")
        else:
            sys.stderr.write(body)
            self._can_rewrite = True
        sys.stderr.flush()

    def finish_message(self, message: str) -> None:
        """Replace the active two-row block with one status line (errors)."""
        msg = message.rstrip("\r\n")
        coloured = _colour_status_message(msg)
        if self._can_rewrite:
            sys.stderr.write(f"\033[2A\r{coloured}\033[K\n\033[2K")
            self._can_rewrite = False
        else:
            sys.stderr.write(f"{coloured}\n")
        sys.stderr.flush()

    def dismiss_two_row_progress(self) -> None:
        """Clear the two-row progress UI before the next phase (summaries print once at the end)."""
        if self._can_rewrite:
            # Erase both rows without a stray \\n between them — that used to leave a visible blank line
            # before the next phase redraw (which writes from row 1 again).
            sys.stderr.write(
                "\033[2A\r\033[2K"
                "\033[1B\r\033[2K"
                "\033[1A"
            )
            self._can_rewrite = False
        sys.stderr.flush()

    def close(self) -> None:
        """Ensure stderr ends cleanly after a progress block (non-TTY)."""
        if self._can_rewrite:
            sys.stderr.write("\n")
            sys.stderr.flush()


def _discover_workbooks_in_source_data(source_root: Path) -> list[Path]:
    """Discover OOXML and Excel-family files under source_root; sorted and de-duplicated."""
    seen: set[Path] = set()
    out: list[Path] = []
    for pattern in ("**/*.xlsx", "**/*.xlsm", "**/*.xltx", "**/*.xltm", "**/*.xls"):
        for p in source_root.glob(pattern):
            if p.is_file():
                r = p.resolve()
                if r not in seen:
                    seen.add(r)
                    out.append(r)
    return sorted(out, key=lambda p: p.as_posix().lower())


def _resolve_workbook_path(raw: Path, script_dir: Path, source_root: Path) -> Path:
    """
    Workbooks must lie under source_root. Relative paths are taken relative to source_root;
    a relative path may also begin with ``source_data/`` or ``data/input/source_data/`` (legacy).
    """
    root = source_root.resolve()
    if raw.is_absolute():
        resolved = raw.resolve()
    else:
        parts = raw.parts
        if parts and parts[0] == "source_data":
            resolved = (DATA_INPUT_DIR / raw).resolve()
        elif len(parts) >= 3 and parts[0] == "data" and parts[1] == "input" and parts[2] == "source_data":
            resolved = (SIM_ROOT / raw).resolve()
        elif len(parts) >= 2 and parts[0] == "data" and parts[1] == "source_data":
            resolved = (DATA_INPUT_DIR / Path(*parts[1:])).resolve()
        else:
            resolved = (source_root / raw).resolve()
    try:
        resolved.relative_to(root)
    except ValueError:
        _stderr_line(f"Error: workbook path must be inside {root}: {raw}")
        raise SystemExit(2)
    return resolved


_NZ_STAMP_MONTHS_EN: tuple[str, ...] = (
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)


def _default_output_csv_path() -> Path:
    """Default CSV path: ``traffic_DDMMMYYYY_hhmmss.csv`` (NZ local time, in ``data/output/demand/``)."""
    dt = datetime.now(ZoneInfo("Pacific/Auckland"))
    mmm = _NZ_STAMP_MONTHS_EN[dt.month - 1].upper()
    stamp = f"{dt.day:02d}{mmm}{dt.year}_{dt.strftime('%H%M%S')}"
    _DEMAND_DIR.mkdir(parents=True, exist_ok=True)
    return _DEMAND_DIR / f"traffic_{stamp}.csv"


def _daily_log_path(*, parent: Path | None = None) -> Path:
    """Daily run log path: ``traffic_DDMMMYYYY.log`` (NZ local date, default: ``data/output/demand/``)."""
    dt = datetime.now(ZoneInfo("Pacific/Auckland"))
    mmm = _NZ_STAMP_MONTHS_EN[dt.month - 1].upper()
    stamp = f"{dt.day:02d}{mmm}{dt.year}"
    base = parent if parent is not None else _DEMAND_DIR
    base.mkdir(parents=True, exist_ok=True)
    return base / f"traffic_{stamp}.log"


def load_street_table(
    path: Path,
) -> dict[str, list[tuple[str, str, str, str]]]:
    """Load CSV with id/Code/intersection_id and street_1…street_4 (case-insensitive headers)."""
    out: dict[str, list[tuple[str, str, str, str]]] = defaultdict(list)
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        # Normalise headers for flexible column names.
        def field(row: dict[str, Any], *names: str) -> str:
            lower = {k.lower().strip(): k for k in row}
            for n in names:
                k = lower.get(n.lower())
                if k is not None:
                    return str(row.get(k) or "").strip()
            return ""

        for row in reader:
            raw = field(row, "id", "Code", "intersection_id").upper()
            if not raw:
                continue
            tup = (
                field(row, "Street_1", "street_1"),
                field(row, "Street_2", "street_2"),
                field(row, "Street_3", "street_3"),
                field(row, "Street_4", "street_4"),
            )
            out[raw].append(tup)
    return dict(out)


def load_intersection_coord_table(path: Path) -> dict[str, dict[str, Any]]:
    """
    Load intersection coordinates from `intersection_geo.csv`.

    Expected headers are case-insensitive. Coordinates are kept as floats when present,
    otherwise empty strings are used so the output CSV has blank cells.
    """
    out: dict[str, dict[str, Any]] = {}
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)

        def field(row: dict[str, Any], *names: str) -> str:
            lower = {k.lower().strip(): k for k in row}
            for n in names:
                k = lower.get(n.lower())
                if k is not None:
                    return str(row.get(k) or "").strip()
            return ""

        def fnum(s: str) -> float | str:
            t = (s or "").strip()
            if not t:
                return ""
            try:
                return float(t)
            except ValueError:
                return ""

        for row in reader:
            iid = field(row, "intersection_id", "id", "Code").upper()
            if not iid:
                continue
            out[iid] = {
                "latitude": fnum(field(row, "latitude", "lat")),
                "longitude": fnum(field(row, "longitude", "lon", "lng")),
            }
    return out


# Octant neighbour columns on intersection_geo.csv (heading N=0°, clockwise).
_NEIGHBOUR_COL_BY_SECTOR_IDX: tuple[str, ...] = (
    "intersection_id_N",
    "intersection_id_NE",
    "intersection_id_E",
    "intersection_id_SE",
    "intersection_id_S",
    "intersection_id_SW",
    "intersection_id_W",
    "intersection_id_NW",
)

# Turning model: inbound travel heading (deg) + delta => outbound heading (then octant).
_MOVEMENT_DELTA_DEG: dict[str, float] = {
    "Thru": 180.0,
    "Left": -90.0,
    "Hard Left": -135.0,
    "Bear Left": -45.0,
    "Right": 90.0,
    "Hard Right": 135.0,
    "Bear Right": 45.0,
    "U-Turn": 360.0,  # (+360) % 360 == 0; keeps outbound sector aligned with inbound
}


def load_intersection_geo_full(path: Path) -> dict[str, dict[str, Any]]:
    """
    Load full intersection_geo rows (coordinates plus directional neighbour ids).

    Neighbour columns are intended to be maintained with
    ``fill_intersection_direction_neighbours.py`` (default **road_chain**: consecutive on a
    named corridor only when segment length ≤ ``--max-edge-m``, then mapped to octants).
    Used to attach destination_id and destination lat/lon from octant neighbours.
    """
    out: dict[str, dict[str, Any]] = {}
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)

        def field(row: dict[str, Any], *names: str) -> str:
            lower = {k.lower().strip(): k for k in row}
            for n in names:
                k = lower.get(n.lower())
                if k is not None:
                    return str(row.get(k) or "").strip()
            return ""

        def fnum(s: str) -> float | str:
            t = (s or "").strip()
            if not t:
                return ""
            try:
                return float(t)
            except ValueError:
                return ""

        for row in reader:
            iid = field(row, "intersection_id", "id", "Code").upper()
            if not iid:
                continue
            rec: dict[str, Any] = {
                "latitude": fnum(field(row, "latitude", "lat")),
                "longitude": fnum(field(row, "longitude", "lon", "lng")),
            }
            for nk in _NEIGHBOUR_COL_BY_SECTOR_IDX:
                rec[nk] = field(row, nk)
            out[iid] = rec
    return out


def _heading_deg_from_bound_label(label: str) -> float | None:
    """Map labels like 'Southbound' / 'Southwestbound' to travel heading (deg from north)."""
    s = _norm_str(label).lower().replace(" ", "")
    if not s or s in ("n/a", "na", "-"):
        return None
    table: dict[str, float] = {
        "northbound": 0.0,
        "northeastbound": 45.0,
        "eastbound": 90.0,
        "southeastbound": 135.0,
        "southbound": 180.0,
        "southwestbound": 225.0,
        "westbound": 270.0,
        "northwestbound": 315.0,
    }
    v = table.get(s)
    return v


def _destination_fields_from_geo_neighbours(
    intersection_id: str,
    approach_bound: str,
    approach_fallback: str,
    movement: str,
    geo_full: dict[str, dict[str, Any]],
) -> tuple[str, Any, Any]:
    """Resolve destination intersection id and coords using neighbour octants."""
    iid = intersection_id.strip().upper()
    g = geo_full.get(iid)
    if not g:
        return ("", "", "")
    inbound = _heading_deg_from_bound_label(approach_bound)
    if inbound is None:
        inbound = _heading_deg_from_bound_label(approach_fallback)
    if inbound is None:
        return ("", "", "")
    delta = _MOVEMENT_DELTA_DEG.get(_norm_str(movement))
    if delta is None:
        return ("", "", "")
    outbound = (inbound + delta) % 360.0
    sector_idx = int(round(outbound / 45.0)) % 8
    nk = _NEIGHBOUR_COL_BY_SECTOR_IDX[sector_idx]
    dest_raw = str(g.get(nk) or "").strip()
    if not dest_raw:
        return ("", "", "")
    dest_id = dest_raw.upper()
    dg = geo_full.get(dest_id)
    if not dg:
        return (dest_id, "", "")
    return (dest_id, dg.get("latitude", ""), dg.get("longitude", ""))


def _apply_destination_enrichment(
    rows: list[dict[str, Any]],
    geo_full: dict[str, dict[str, Any]],
) -> None:
    """Set destination_id, destination_latitude, destination_longitude on each row (in place)."""
    for row in rows:
        did, dlat, dlon = _destination_fields_from_geo_neighbours(
            str(row.get("intersection_id") or ""),
            str(row.get("approach_bound") or ""),
            str(row.get("approach_header_2") or ""),
            str(row.get("movement") or ""),
            geo_full,
        )
        row["destination_id"] = did
        row["destination_latitude"] = dlat
        row["destination_longitude"] = dlon


def resolve_street_names(
    intersection_id: str,
    workbook_label: str,
    lookup: dict[str, list[tuple[str, str, str, str]]],
) -> tuple[str, str, str, str]:
    """Pick street tuple for ID; disambiguate duplicates using workbook label text."""
    iid = intersection_id.strip().upper()
    opts = lookup.get(iid, [])
    if not opts:
        return ("", "", "", "")
    if len(opts) == 1:
        return opts[0]
    lb = workbook_label.lower()
    for s1, s2, s3, s4 in opts:
        for part in (s1, s2, s3, s4):
            if len(part) > 2 and part.lower() in lb:
                return (s1, s2, s3, s4)
    for s1, s2, s3, s4 in opts:
        if s1 and s1.lower() in lb:
            return (s1, s2, s3, s4)
    return opts[0]


def parse_loaded_workbook(
    wb: Any,
    path: Path,
    *,
    include_metadata: bool = True,
    include_source_file: bool = False,
    street_lookup: dict[str, list[tuple[str, str, str, str]]] | None = None,
    coord_lookup: dict[str, dict[str, Any]] | None = None,
    geo_full: dict[str, dict[str, Any]] | None = None,
    use_workbook_intersection_label: bool = False,
) -> list[dict[str, Any]]:
    workbook_label = _extract_intersection_label(wb, path)
    intersection_id = _compute_intersection_id(path, wb)
    survey_dt = _compute_survey_date(wb, path)
    survey_date_str = survey_dt.isoformat() if survey_dt else ""
    roles = _miovision_metric_sheet_roles(wb)
    pick_k, _pick_ws = _pick_layout_worksheet(wb, roles)
    if pick_k is None:
        raise ValueError(
            "Expected at least one Miovision metric sheet with a movement grid "
            '(column A label "Start Time"). No Lights / Other Vehicles / Bicycles / '
            "Totals sheet matched, or none contained the grid."
        )
    rows = parse_miovision_metric_workbook(wb, roles=roles)
    if not include_metadata:
        return rows
    meta: dict[str, Any] = {"intersection_id": intersection_id, "survey_date": survey_date_str}
    coords = (coord_lookup or {}).get(intersection_id.strip().upper())
    if coords:
        meta.update(coords)
    else:
        meta["latitude"] = ""
        meta["longitude"] = ""
    if use_workbook_intersection_label:
        meta["street_1"] = workbook_label
        meta["street_2"] = ""
        meta["street_3"] = ""
        meta["street_4"] = ""
    else:
        lu = street_lookup or {}
        s1, s2, s3, s4 = resolve_street_names(intersection_id, workbook_label, lu)
        meta["street_1"] = s1
        meta["street_2"] = s2
        meta["street_3"] = s3
        meta["street_4"] = s4
    if include_source_file:
        meta["source_file"] = path.name
    out = [{**meta, **row} for row in rows]
    if geo_full:
        _apply_destination_enrichment(out, geo_full)
    else:
        for r in out:
            r["destination_id"] = ""
            r["destination_latitude"] = ""
            r["destination_longitude"] = ""
    return out


def parse_xlsx(
    path: Path,
    *,
    include_metadata: bool = True,
    include_source_file: bool = False,
    street_lookup: dict[str, list[tuple[str, str, str, str]]] | None = None,
    coord_lookup: dict[str, dict[str, Any]] | None = None,
    geo_full: dict[str, dict[str, Any]] | None = None,
    use_workbook_intersection_label: bool = False,
) -> list[dict[str, Any]]:
    wb = open_workbook_auto(path)
    try:
        return parse_loaded_workbook(
            wb,
            path,
            include_metadata=include_metadata,
            include_source_file=include_source_file,
            street_lookup=street_lookup,
            coord_lookup=coord_lookup,
            geo_full=geo_full,
            use_workbook_intersection_label=use_workbook_intersection_label,
        )
    finally:
        close_workbook(wb)


def write_csv(
    rows: list[dict[str, Any]],
    out_path: Path,
    *,
    include_source_file: bool = False,
) -> None:
    fieldnames = []
    if include_source_file:
        fieldnames.append("source_file")
    fieldnames.extend(
        [
            "intersection_id",
            "latitude",
            "longitude",
            "street_1",
            "street_2",
            "street_3",
            "street_4",
            "survey_date",
            "period",
            "time",
            "movement_index",
            "approach_header_1",
            "approach_header_2",
            "approach_bound",
            "movement",
            "destination_id",
            "destination_latitude",
            "destination_longitude",
            "lights",
            "other_vehicles",
            "bicycles_on_road",
            "bicycles_crosswalk",
            "totals",
        ]
    )
    rows = sorted(
        rows,
        key=lambda r: (r.get("time") or "", r.get("intersection_id") or ""),
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, **_CSV_WRITE_OPTS)
        w.writeheader()
        w.writerows(rows)


def _workbook_matches_scan_record(inp: Path, wb: Any, exp_iid: str, exp_iso: str) -> bool:
    """True when workbook-derived id and survey date match the light scan."""
    wb_iid = _compute_intersection_id(inp, wb)
    wb_dt = _compute_survey_date(wb, inp)
    wb_iso = wb_dt.isoformat() if wb_dt else ""
    return wb_iid == exp_iid and wb_iso == exp_iso


def _collect_scan_metadata(
    all_inputs: list[Path],
    prog: _DualPhaseProgress,
    show_ui: bool,
) -> dict[Path, tuple[str | None, datetime | None]]:
    """
    Phase 1 — SCANNING: intersection id and survey date per source file (open workbook if needed).
    """
    pending: dict[Path, tuple[str | None, datetime | None]] = {}
    n_cat = len(all_inputs)
    for idx, inp in enumerate(all_inputs, start=1):
        if show_ui:
            prog.tick_scan(
                scan_index=idx,
                scan_total=n_cat,
                scan_file=inp.name,
            )
        iid_fn = _extract_intersection_id_from_filename(inp)
        dt_fn = _extract_survey_date_from_filename(inp)
        iid = iid_fn
        survey_dt = dt_fn

        if iid is None or survey_dt is None:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", UserWarning)
                    wb = open_workbook_auto(inp)
                try:
                    if iid is None:
                        iid = _compute_intersection_id(inp, wb)
                    if survey_dt is None:
                        survey_dt = _compute_survey_date(wb, inp)
                finally:
                    close_workbook(wb)
            except Exception as e:
                _plain_mode_file_line(
                    show_ui,
                    ok=False,
                    ok_prefix="SCANNED",
                    workbook=inp,
                    survey_date=survey_dt if isinstance(survey_dt, date) else None,
                )
                if show_ui:
                    prog.finish_message(f"Error ({inp.name}): {e}")
                else:
                    _stderr_line(f"Error ({inp.name}): {e}")
                raise SystemExit(1) from e

        pending[inp] = (iid, survey_dt)
        # TTY: progress bar only on stderr; per-file SCANNED lines go to the run log only (see _plain_mode_file_line).
        _plain_mode_file_line(
            show_ui,
            ok=True,
            ok_prefix="SCANNED",
            workbook=inp,
            survey_date=survey_dt if isinstance(survey_dt, date) else None,
        )

    if show_ui:
        prog.dismiss_two_row_progress()
    return pending


def _apply_filters_and_dedupe(
    all_inputs: list[Path],
    pending: dict[Path, tuple[str | None, datetime | None]],
    *,
    allow_filter: bool,
    allowed_from_csv: frozenset[str],
    year_filter: bool,
    min_survey_year: int,
    prog: _DualPhaseProgress,
    show_ui: bool,
) -> tuple[dict[Path, tuple[str, str]], set[Path]]:
    """
    Phase 2 — VALIDATING: street table / year rules and newest-wins per intersection id.
    """
    best: dict[str, tuple[str, int, Path]] = {}
    scan_ok: dict[Path, tuple[str, str]] = {}
    n_cat = len(all_inputs)
    for idx, inp in enumerate(all_inputs, start=1):
        if show_ui:
            prog.tick_validate(
                validate_index=idx,
                validate_total=n_cat,
                validate_file=inp.name,
            )
        iid, survey_dt = pending[inp]

        if allow_filter:
            if not iid:
                _plain_mode_file_line(
                    show_ui,
                    ok=False,
                    ok_prefix="VALIDATED",
                    workbook=inp,
                    survey_date=survey_dt if isinstance(survey_dt, date) else None,
                )
                continue
            if allowed_from_csv and iid not in allowed_from_csv:
                _plain_mode_file_line(
                    show_ui,
                    ok=False,
                    ok_prefix="VALIDATED",
                    workbook=inp,
                    survey_date=survey_dt if isinstance(survey_dt, date) else None,
                )
                continue
        if year_filter:
            if survey_dt is None:
                _plain_mode_file_line(
                    show_ui,
                    ok=False,
                    ok_prefix="VALIDATED",
                    workbook=inp,
                    survey_date=None,
                )
                continue
            if survey_dt.year < min_survey_year:
                _plain_mode_file_line(
                    show_ui,
                    ok=False,
                    ok_prefix="VALIDATED",
                    workbook=inp,
                    survey_date=survey_dt if isinstance(survey_dt, date) else None,
                )
                continue

        survey_iso = survey_dt.isoformat() if survey_dt else ""
        scan_ok[inp] = (iid, survey_iso)

        cur = best.get(iid)
        if cur is None:
            best[iid] = (survey_iso, idx, inp)
        else:
            old_iso, old_idx, old_path = cur
            if (survey_iso, idx) >= (old_iso, old_idx):
                best[iid] = (survey_iso, idx, inp)
            else:
                best[iid] = (old_iso, old_idx, old_path)

    winners = {p for (_iso, _idx, p) in best.values()}
    # Always record dedupe outcomes in the run log; non-TTY also prints them to stderr.
    for inp in all_inputs:
        if inp not in scan_ok:
            continue
        if inp in winners:
            _plain_mode_file_line(
                show_ui,
                ok=True,
                ok_prefix="VALIDATED",
                workbook=inp,
                survey_date=_iso_survey_date(scan_ok[inp][1]),
            )
        else:
            _plain_mode_file_line(
                show_ui,
                ok=False,
                ok_prefix="VALIDATED",
                workbook=inp,
                survey_date=_iso_survey_date(scan_ok[inp][1]),
            )

    if show_ui:
        prog.dismiss_two_row_progress()
    return scan_ok, winners


def _verify_winners_open_workbooks(
    winner_paths: list[Path],
    scan_ok: dict[Path, tuple[str, str]],
    prog: _DualPhaseProgress,
    show_ui: bool,
) -> list[Path]:
    """Open each winner workbook and confirm id/date match the scan; return paths kept."""
    n_winners = len(winner_paths)
    verified_winners: list[Path] = []
    for wi, inp in enumerate(winner_paths, start=1):
        if show_ui:
            prog.tick_verify(
                verify_index=wi,
                verify_total=n_winners,
                verify_file=inp.name,
            )
        exp_iid, exp_iso = scan_ok[inp]
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                wb = open_workbook_auto(inp)
            try:
                if not _workbook_matches_scan_record(inp, wb, exp_iid, exp_iso):
                    _plain_mode_file_line(
                        show_ui,
                        ok=False,
                        ok_prefix="VERIFYING",
                        workbook=inp,
                        survey_date=_iso_survey_date(exp_iso),
                    )
                    continue
            finally:
                close_workbook(wb)
        except Exception as e:
            _plain_mode_file_line(
                show_ui,
                ok=False,
                ok_prefix="VERIFYING",
                workbook=inp,
                survey_date=_iso_survey_date(exp_iso),
            )
            if show_ui:
                prog.finish_message(f"Error ({inp.name}): {e}")
            else:
                _stderr_line(f"Error ({inp.name}): {e}")
            raise SystemExit(1) from e

        _plain_mode_file_line(
            show_ui,
            ok=True,
            ok_prefix="VERIFIED",
            workbook=inp,
            survey_date=_iso_survey_date(exp_iso),
        )
        verified_winners.append(inp)
    return verified_winners


def _parse_verified_workbooks(
    verified_winners: list[Path],
    *,
    meta: bool,
    street_lookup: dict[str, list[tuple[str, str, str, str]]],
    coord_lookup: dict[str, dict[str, Any]],
    geo_full: dict[str, dict[str, Any]],
    prog: _DualPhaseProgress,
    show_ui: bool,
    survey_date_by_path: dict[Path, date | None] | None = None,
) -> list[dict[str, Any]]:
    """Parse each verified workbook into rows (flat list in intersection order)."""
    buckets: dict[str, list[dict[str, Any]]] = {}
    label_order: list[str] = []
    rows_so_far = 0
    parse_total_w = max(1, len(verified_winners))

    dates = survey_date_by_path or {}
    for pi, inp in enumerate(verified_winners, start=1):
        chunk: list[dict[str, Any]] | None = None
        sd_parse = dates.get(inp)
        if show_ui:
            prog.tick_parse(
                parse_index=pi,
                parse_total=parse_total_w,
                parse_file=inp.name,
                rows_so_far=rows_so_far,
            )
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                wb = open_workbook_auto(inp)
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", UserWarning)
                    chunk = parse_loaded_workbook(
                        wb,
                        inp,
                        include_metadata=meta,
                        include_source_file=False,
                        street_lookup=street_lookup,
                        coord_lookup=coord_lookup,
                        geo_full=geo_full,
                        use_workbook_intersection_label=False,
                    )
            finally:
                close_workbook(wb)
        except Exception as e:
            _plain_mode_file_line(
                show_ui,
                ok=False,
                ok_prefix="PARSING",
                workbook=inp,
                survey_date=sd_parse,
            )
            if show_ui:
                prog.finish_message(f"Error ({inp.name}): {e}")
            else:
                _stderr_line(f"Error ({inp.name}): {e}")
            raise SystemExit(1) from e

        if not chunk:
            _plain_mode_file_line(
                show_ui,
                ok=False,
                ok_prefix="PARSING",
                workbook=inp,
                survey_date=sd_parse,
            )
            continue

        key = chunk[0].get("intersection_id") or inp.stem
        if key not in buckets:
            label_order.append(key)
            buckets[key] = chunk
        else:
            buckets[key] = chunk
        rows_so_far += len(chunk)
        _plain_mode_file_line(
            show_ui,
            ok=True,
            ok_prefix="PARSED",
            workbook=inp,
            survey_date=sd_parse,
        )

    all_rows: list[dict[str, Any]] = []
    for k in label_order:
        all_rows.extend(buckets[k])
    return all_rows


def main() -> None:
    # Non-TTY stderr is often buffered by wrappers; line-buffering keeps logs readable.
    if not sys.stderr.isatty():
        try:
            sys.stderr.reconfigure(line_buffering=True)
        except (AttributeError, OSError, ValueError):
            pass

    ap = argparse.ArgumentParser(
        prog=Path(__file__).name,
        description=(
            "Convert Miovision-style traffic-count Excel workbooks (metric sheets such as "
            "Lights, Other Vehicles, Bicycles, Totals — any subset) to long CSV. "
            "Workbooks must live under data/input/source_data/."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Usage:\n"
            "  Run everything under data/input/source_data/:\n"
            "    python3 traffic_counts_parser.py\n"
            "\n"
            "  Run specific workbooks (paths are under data/input/source_data/):\n"
            "    python3 traffic_counts_parser.py \"2020 Intersection/file.xlsx\"\n"
            "\n"
            "References:\n"
            "  Christchurch City Council (intersection traffic counts database):\n"
            "    https://ccc.govt.nz/transport/improving-our-transport-and-roads/traffic-count-data/intersection-traffic-counts-database\n"
            "  Intersection workbook folder (Google Drive):\n"
            "    https://drive.google.com/drive/folders/1oP5gcuKR1bHB9Xn2B4ILZZd_LhpMggxU\n"
        ),
    )
    ap.add_argument(
        "inputs",
        type=Path,
        nargs="*",
        default=[],
        help=(
            "Paths under data/input/source_data/ (e.g. '2020 Intersection/file.xlsx'). "
            "If omitted, all supported workbook types under data/input/source_data are used "
            "(.xlsx, .xlsm, .xltx, .xltm, .xls)."
        ),
    )
    args = ap.parse_args()

    script_dir = Path(__file__).resolve().parent
    source_root = DATA_INPUT_DIR / "source_data"
    if not source_root.is_dir():
        _stderr_line(f"Error: source_data folder not found: {source_root}")
        raise SystemExit(2)

    raw_list: list[Path] = list(args.inputs)
    if not raw_list:
        all_inputs = _discover_workbooks_in_source_data(source_root)
        if not all_inputs:
            _stderr_line(
                "Error: no supported workbook files (.xlsx, .xlsm, .xltx, .xltm, .xls) "
                f"under {source_root.resolve()}"
            )
            raise SystemExit(2)
    else:
        all_inputs = [
            _resolve_workbook_path(p, script_dir, source_root) for p in raw_list
        ]
    streets_path = DATA_OUTPUT_DIR / "intersection_geo.csv"
    street_lookup: dict[str, list[tuple[str, str, str, str]]] = {}
    coord_lookup: dict[str, dict[str, Any]] = {}
    geo_full: dict[str, dict[str, Any]] = {}
    if streets_path.is_file():
        street_lookup = load_street_table(streets_path)
        coord_lookup = load_intersection_coord_table(streets_path)
        geo_full = load_intersection_geo_full(streets_path)
    else:
        _stderr_line(
            f"Warning: intersection geo table not found at {streets_path}; "
            "street_1…street_4 will be empty and intersection ID filtering is disabled."
        )
    allowed_from_csv: frozenset[str] = frozenset(street_lookup.keys())
    if not allowed_from_csv:
        _stderr_line(
            "Warning: no intersection ids in the street table; "
            "intersection ID filter is not applied."
        )

    for p in all_inputs:
        if not p.is_file():
            _stderr_line(f"Not found: {p}")
            raise SystemExit(2)

    out = _default_output_csv_path()
    log_path = _daily_log_path(parent=out.parent)
    global _RUN_LOG_FH
    _RUN_LOG_FH = log_path.open("a", encoding="utf-8", newline="", buffering=1)

    meta = True
    allow_filter = True
    year_filter = True
    min_survey_year = 2016

    prog = _DualPhaseProgress()
    show_ui = sys.stderr.isatty()
    if show_ui:
        _stderr_line("")

    pending = _collect_scan_metadata(all_inputs, prog, show_ui)
    if not show_ui:
        _stderr_line("")
    else:
        _emit_tty_summary_line("SCANNED:", len(all_inputs), "files")
    if show_ui:
        prog.invalidate()
    scan_ok, winners = _apply_filters_and_dedupe(
        all_inputs,
        pending,
        allow_filter=allow_filter,
        allowed_from_csv=allowed_from_csv,
        year_filter=year_filter,
        min_survey_year=min_survey_year,
        prog=prog,
        show_ui=show_ui,
    )

    if not show_ui:
        _stderr_line("")
    else:
        _emit_tty_summary_line("VALIDATED:", len(scan_ok), "files")
    if show_ui:
        prog.invalidate()

    winner_paths = [p for p in all_inputs if p in winners]
    verified_winners = _verify_winners_open_workbooks(
        winner_paths,
        scan_ok,
        prog,
        show_ui,
    )

    if show_ui:
        prog.dismiss_two_row_progress()
        _emit_tty_summary_line("VERIFIED:", len(verified_winners), "files")

    if not show_ui:
        _stderr_line("")

    if verified_winners:
        prog.begin_parse_phase()

    n_parse = len(verified_winners)
    parse_survey_dates = {
        p: _iso_survey_date(scan_ok[p][1]) for p in verified_winners
    }
    all_rows = _parse_verified_workbooks(
        verified_winners,
        meta=meta,
        street_lookup=street_lookup,
        coord_lookup=coord_lookup,
        geo_full=geo_full,
        prog=prog,
        show_ui=show_ui,
        survey_date_by_path=parse_survey_dates,
    )

    if not all_rows:
        _stderr_line("Warning: no data rows written.")

    write_csv(
        all_rows,
        out,
        include_source_file=False,
    )
    try:
        if show_ui:
            prog.dismiss_two_row_progress()
        else:
            prog.close()
        if show_ui:
            _emit_tty_summary_line("PARSED:", len(all_rows), "rows", num_width=7)
            _stderr_line("")
            _stderr_line(_colour_status_message(f"COMPLETED: {out.name}"))
            _stderr_line("")
        else:
            _stderr_line("")
            _emit_final_summary_block(
                tty_colours=False,
                n_scan=len(all_inputs),
                n_val=len(scan_ok),
                n_ver=len(verified_winners),
                n_row=len(all_rows),
            )
            _stderr_line("")
            _stderr_line(_colour_status_message(f"COMPLETED: {out.name}"))
            _stderr_line("")
    finally:
        if _RUN_LOG_FH is not None:
            _RUN_LOG_FH.close()
            _RUN_LOG_FH = None


if __name__ == "__main__":
    main()
